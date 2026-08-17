#!/usr/bin/env python3
"""
generate_fixtures_page.py
Reads Fixtures_2 from fingallians_results.xlsx, builds a styled HTML fixtures
table (today onwards only), and posts it to the Fixtures & Results page on t
fingallians.ie via the WordPress REST API.

Usage:
    python3 generate_fixtures_page.py

Requirements:
    pip install openpyxl requests
"""

import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import requests
from requests.auth import HTTPBasicAuth

# ── Config ────────────────────────────────────────────────────────────────────

XLSX_PATH   = Path(__file__).parent / "fingallians_results.xlsx"
SHEET_NAME  = "Fixtures_2"

WP_BASE_URL = "https://fingallians.ie/wp-json/wp/v2"
WP_PAGE_ID  = 127          # "Fixtures & Results" page
WP_USER     = "fins_admin"
WP_APP_PASS = "6u9w x2Z0 aPqJ ADSm PWSn rDRy"
                           #   (Dashboard → Users → Profile → Application Passwords)

# ── Date parsing ──────────────────────────────────────────────────────────────

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
}

DAY_NAMES = {"monday", "tuesday", "wednesday", "thursday",
             "friday", "saturday", "sunday"}

CURRENT_YEAR = date.today().year


def parse_date(s: str) -> date | None:
    """Parse date strings in two formats used by Sportlomo / the scraper.

    Format A (newer manual fixtures): '13 June 2026'
    Format B (older API format):      'Saturday August 15th'
    """
    if not s:
        return None
    s = s.strip()

    # Format A: '13 June 2026'
    try:
        return datetime.strptime(s, "%d %B %Y").date()
    except ValueError:
        pass

    # Format B: 'Saturday August 15th'  (ordinal suffix, no year)
    parts = s.split()
    if len(parts) >= 3 and parts[0].lower() in DAY_NAMES:
        month_str = parts[1].lower()
        day_str   = re.sub(r"(st|nd|rd|th)$", "", parts[2], flags=re.I)
        month_num = MONTH_MAP.get(month_str)
        if month_num and day_str.isdigit():
            try:
                return date(CURRENT_YEAR, month_num, int(day_str))
            except ValueError:
                pass

    return None


# ── Competition short-form labels ─────────────────────────────────────────────

def shorten_competition(comp: str) -> str:
    comp = comp.strip()
    # Adult Football League
    m = re.match(r"PTSB Adult Football League\s+(\d+\w*)", comp, re.I)
    if m:
        return f"AFL Div {m.group(1)}"
    # Adult Hurling League
    m = re.match(r"PTSB Adult Hurling League\s+(\d+\w*)", comp, re.I)
    if m:
        return f"AHL Div {m.group(1)}"
    # Minor Football League
    m = re.match(r"PTSB Minor Football League Division\s+(\d+\w*)", comp, re.I)
    if m:
        return f"MinFL Div {m.group(1)}"
    # Minor Hurling League
    m = re.match(r"PTSB Minor Hurling League Division\s+(\d+\w*)", comp, re.I)
    if m:
        return f"MinHL Div {m.group(1)}"
    # Premier IHC
    if re.search(r"Premier Intermediate Hurling Championship", comp, re.I):
        return "Premier IHC"
    # Go Games / Juvenile – strip region suffix
    comp = re.sub(r"\s*\(\d+ team\)[^$]*$", "", comp).strip()
    return comp


# ── Sport / category detection ────────────────────────────────────────────────

def detect_category(comp: str) -> str:
    c = comp.lower()
    if re.search(r"premier intermediate hurling championship", c):
        return "Adult Hurling Championship"
    if re.search(r"adult hurling league", c):
        return "Adult Hurling"
    if re.search(r"adult football league", c):
        return "Adult Football"
    if re.search(r"minor hurling league", c):
        return "Minor Hurling"
    if re.search(r"minor football league", c):
        return "Minor Football"
    if re.search(r"u1[5-9]|u2[0-9]", c) and "football" in c:
        return "Minor Football"
    if re.search(r"u1[5-9]|u2[0-9]", c) and "hurling" in c:
        return "Minor Hurling"
    # Go Games / juvenile
    return "Go Games Football"


# ── Format date for display ───────────────────────────────────────────────────

def ordinal(n: int) -> str:
    suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10 if n % 100 not in (11, 12, 13) else 0, "th")
    return f"{n}{suffix}"


def format_date(d: date) -> str:
    return f"{d.strftime('%A')} {d.strftime('%B')} {ordinal(d.day)}"


# ── Read & filter fixtures ────────────────────────────────────────────────────

def load_fixtures(today: date) -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    fixtures = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        r = dict(zip(headers, row))
        d = parse_date(str(r.get("Date", "") or ""))
        if d is None or d < today:
            continue
        r["_date"]   = d
        r["_label"]  = format_date(d)
        r["_short"]  = shorten_competition(str(r.get("Competition", "") or ""))
        r["_cat"]    = detect_category(str(r.get("Competition", "") or ""))
        fixtures.append(r)

    fixtures.sort(key=lambda x: (x["_date"], x["_cat"], x.get("Time") or ""))
    return fixtures


# ── HTML generation ───────────────────────────────────────────────────────────

CSS = """<style>
.fins-fixtures-wrap { font-family: Arial, Helvetica, sans-serif; font-size: 14px; margin: 20px 0; }
.fins-fixtures-wrap h2 { text-align: center; font-size: 22px; color: #8B0000; margin-bottom: 18px; letter-spacing: 0.5px; }
.fins-fixtures-table { width: 100%; border-collapse: collapse; }
.fins-fixtures-table .date-row td { background: #FFE8E8; color: #8B0000; font-weight: bold; font-size: 16px; text-align: center; padding: 10px 8px; border-top: 3px solid #8B0000; }
.fins-fixtures-table .cat-row td { background: #8B0000; color: #fff; font-weight: bold; text-align: center; padding: 7px 8px; font-size: 13px; letter-spacing: 1px; text-transform: uppercase; }
.fins-fixtures-table .fix-row td { padding: 8px 10px; border-bottom: 1px solid #eedddd; vertical-align: middle; }
.fins-fixtures-table .fix-row:nth-child(even) { background: #FFF5F5; }
.fins-fixtures-table .fix-row:nth-child(odd) { background: #ffffff; }
.fins-fixtures-table .col-comp { width: 20%; }
.fins-fixtures-table .col-venue { width: 20%; }
.fins-fixtures-table .col-match { width: 35%; font-weight: 500; }
.fins-fixtures-table .col-time { width: 8%; text-align: center; }
.fins-fixtures-table .col-ground { width: 10%; text-align: center; }
.badge-home { display: inline-block; background: #1a7a3a; color: #fff; border-radius: 3px; padding: 2px 7px; font-size: 11px; font-weight: bold; }
.badge-away { display: inline-block; background: #c47a00; color: #fff; border-radius: 3px; padding: 2px 7px; font-size: 11px; font-weight: bold; }
@media (max-width: 600px) {
  .fins-fixtures-table .col-comp, .fins-fixtures-table .col-venue { display: none; }
  .fins-fixtures-table .col-match { width: 60%; }
}
</style>"""


def build_html(fixtures: list[dict]) -> str:
    if not fixtures:
        return "<!-- wp:html -->\n<p>No upcoming fixtures.</p>\n<!-- /wp:html -->"

    rows = []
    last_date = None
    last_cat  = None

    for fx in fixtures:
        # Date header
        if fx["_label"] != last_date:
            rows.append(f'<tr class="date-row"><td colspan="5">{fx["_label"]}</td></tr>')
            last_date = fx["_label"]
            last_cat  = None

        # Category header
        if fx["_cat"] != last_cat:
            rows.append(f'<tr class="cat-row"><td colspan="5">{fx["_cat"]}</td></tr>')
            last_cat = fx["_cat"]

        # Fixture row
        home   = str(fx.get("Your Club Name") or "Fingallians")
        opp    = str(fx.get("Opponent") or "")
        ground = str(fx.get("Ground") or "").strip().lower()
        venue  = str(fx.get("Venue") or "")
        time_  = str(fx.get("Time") or "")
        comp   = fx["_short"]

        if ground == "home":
            match_str = f"{home} v {opp}"
            badge     = '<span class="badge-home">Home</span>'
        else:
            match_str = f"{opp} v {home}"
            badge     = '<span class="badge-away">Away</span>'

        rows.append(
            f'<tr class="fix-row">'
            f'<td class="col-comp">{comp}</td>'
            f'<td class="col-venue">{venue}</td>'
            f'<td class="col-match">{match_str}</td>'
            f'<td class="col-time">{time_}</td>'
            f'<td class="col-ground">{badge}</td>'
            f'</tr>'
        )

    table = "\n".join(rows)
    html = (
        f"<!-- wp:html -->\n"
        f"{CSS}\n"
        f'<div class="fins-fixtures-wrap">\n'
        f"<h2>Fingallians Upcoming Fixtures</h2>\n"
        f'<table class="fins-fixtures-table">\n'
        f"{table}\n"
        f"</table></div>\n"
        f"<!-- /wp:html -->"
    )
    return html


# ── WordPress update ──────────────────────────────────────────────────────────

def get_page_content(session: requests.Session) -> str:
    """Fetch the current page content, return everything before our fixtures block."""
    resp = session.get(
        f"{WP_BASE_URL}/pages/{WP_PAGE_ID}",
        params={"context": "edit"},
    )
    resp.raise_for_status()
    raw = resp.json()["content"]["raw"]

    # Strip any previously injected fixtures block (everything from <!-- wp:html --> onwards
    # that contains our fins-fixtures marker)
    marker = "<!-- wp:html -->"
    idx = raw.find(marker)
    if idx != -1 and "fins-fixtures" in raw[idx:]:
        raw = raw[:idx].rstrip()
    return raw


def post_to_wordpress(fixtures_html: str) -> None:
    if not WP_APP_PASS:
        print("\n⚠️  WP_APP_PASS is not set.")
        print("   Go to: Dashboard → Users → Profile → Application Passwords")
        print("   Generate a password, paste it into WP_APP_PASS in this script, then re-run.")
        sys.exit(1)

    session = requests.Session()
    session.auth = HTTPBasicAuth(WP_USER, WP_APP_PASS)

    print("Fetching current page content …")
    base_content = get_page_content(session)

    new_content = base_content + "\n\n" + fixtures_html if base_content else fixtures_html

    print("Posting updated content …")
    resp = session.post(
        f"{WP_BASE_URL}/pages/{WP_PAGE_ID}",
        json={"content": new_content},
    )
    resp.raise_for_status()
    data = resp.json()
    print(f"✅  Page updated — ID {data['id']}, modified {data['modified']}")
    print(f"   View at: https://fingallians.ie/fixtures-results/")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    today = date.today()
    print(f"Generating fixtures from {today} onwards …")

    fixtures = load_fixtures(today)
    print(f"Found {len(fixtures)} upcoming fixtures.")

    html = build_html(fixtures)
    print(f"Generated {len(html):,} chars of HTML.")

    post_to_wordpress(html)


if __name__ == "__main__":
    main()
