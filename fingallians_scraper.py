"""
Fingallians GAA Fixtures & Results Scraper

First run:  fetches all results from 1 Jan of the current year and creates
            fingallians_results.xlsx

Later runs: fetches only results since the last saved date and appends new
            rows — no duplicates.

Usage:
    pip install requests beautifulsoup4 openpyxl
    python3 fingallians_scraper.py

Options (edit below):
    SPORT_FILTER  — "football", "hurling", "all"
"""

import re
import sys
from datetime import datetime, date
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ─────────────────────────────────────────────────────────────
CLUB_ID    = "3573"
API_URL    = "https://dublingaa.sportlomo.com/wp-admin/admin-ajax.php"
OUTPUT_FILE = Path(__file__).parent / "fingallians_results.xlsx"

# Set to "football", "hurling", or "all"
SPORT_FILTER = "all"

# Each entry is (org, user_id) — add extra rows here if more boards are found.
# NOTE: org=3 covers adult + minor (U16/U17+) competitions.
# Go Games / U12 and younger are NOT exposed via the club_page_fixtures API;
# they live on individual competition pages (/league-2/XXXXX) and would need
# separate scraping if required.
ORGS = [
    ("3", "3"),  # Adult / Senior / Minor — Dublin GAA main board
]

COLUMNS = ["Date", "Sport", "Competition", "Time",
           "Home Team", "Home Score", "Away Team", "Away Score",
           "Result", "Venue", "Notes"]
COL_WIDTHS = {
    "Date": 18, "Sport": 12, "Competition": 30, "Time": 8,
    "Home Team": 26, "Home Score": 11, "Away Team": 26,
    "Away Score": 11, "Result": 8, "Venue": 22, "Notes": 28,
}

COLUMNS_F2 = ["Date", "Time", "Venue", "Ground", "Referee", "Team",
              "Competition", "Your Club Name", "Opponent", "Event Type"]
COL_WIDTHS_F2 = {
    "Date": 18, "Time": 8, "Venue": 22, "Ground": 8, "Referee": 16,
    "Team": 30, "Competition": 30, "Your Club Name": 16,
    "Opponent": 26, "Event Type": 14,
}

MONTH_MAP = {m.lower(): i for i, m in enumerate(
    ["January","February","March","April","May","June",
     "July","August","September","October","November","December"], 1)}

FOOTBALL_KEYWORDS = ["football", "football league", "football championship"]
HURLING_KEYWORDS  = ["hurling", "camogie"]


# ── Date helpers ───────────────────────────────────────────────────────────────

def parse_date(text):
    text = re.sub(r"(st|nd|rd|th)", "", text, flags=re.I)
    text = re.sub(r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s*", "", text, flags=re.I)
    text = text.strip()
    for month_name, month_num in MONTH_MAP.items():
        if month_name in text.lower():
            m = re.search(r"(\d+)", text)
            if m:
                day = int(m.group(1))
                year = date.today().year
                try:
                    d = date(year, month_num, day)
                    if (d - date.today()).days > 180:
                        d = date(year - 1, month_num, day)
                    return d
                except ValueError:
                    pass
    return None


def date_to_str(d):
    return d.strftime("%-d %B %Y") if d else ""


def detect_sport(competition):
    comp_lower = competition.lower()
    if any(k in comp_lower for k in FOOTBALL_KEYWORDS):
        return "Football"
    if any(k in comp_lower for k in HURLING_KEYWORDS):
        return "Hurling"
    return "Other"


def build_fixtures2(fixtures):
    """Transform raw fixtures list into the Fixtures_2 column format."""
    rows = []
    for f in fixtures:
        home = f.get("Home Team", "")
        away = f.get("Away Team", "")
        competition = f.get("Competition", "")

        is_home  = "Fingallians" in home
        ground   = "Home" if is_home else "Away"
        opponent = away if is_home else home

        comp_lower = competition.lower()
        if "championship" in comp_lower:
            event_type = "Championship"
        else:
            event_type = "League"

        rows.append({
            "Date":          f.get("Date", ""),
            "Time":          f.get("Time", ""),
            "Venue":         f.get("Venue", ""),
            "Ground":        ground,
            "Referee":       "",
            "Team":          competition,
            "Competition":   competition,
            "Your Club Name": "Fingallians",
            "Opponent":      opponent,
            "Event Type":    event_type,
        })
    return rows


def parse_manual_fixtures_file():
    """Parse a Manual_Fixtures text file (raw copy-paste from the Sportlomo page).

    The file lives in the same directory as this script. Each fixture block
    produced by Sportlomo repeats team names, venue, and referee twice (once for
    desktop, once for mobile), with a 'GAA Logo GAA Logo' line between the away
    team and the venue.  The parser handles all of that automatically.

    The file can contain any fixtures — duplicates are removed by the dedup()
    call in main(), so it is fine to paste the full page output including adult
    and minor fixtures that the API already returns.
    """
    filepath = Path(__file__).parent / "Manual_Fixtures"
    if not filepath.exists():
        return []

    DAY_NAMES   = {"monday", "tuesday", "wednesday", "thursday",
                   "friday", "saturday", "sunday"}
    MONTH_NAMES = set(MONTH_MAP.keys())

    # Read non-empty lines, stripping whitespace
    lines = []
    with open(filepath, encoding="utf-8") as fh:
        for line in fh:
            s = line.strip()
            if s:
                lines.append(s)

    def is_date_line(s):
        parts = s.lower().split()
        return (len(parts) >= 3
                and parts[0] in DAY_NAMES
                and parts[1] in MONTH_NAMES)

    def is_time_line(s):
        return bool(re.match(r"^\d{1,2}:\d{2}$", s))

    fixtures = []
    i = 0
    while i < len(lines):
        # Skip non-date lines (header, filter UI text, etc.)
        if not is_date_line(lines[i]):
            i += 1
            continue

        date_str    = lines[i]; i += 1
        if i >= len(lines): break

        competition = lines[i]; i += 1
        if i >= len(lines) or not is_time_line(lines[i]):
            continue

        time_str    = lines[i]; i += 1
        if i >= len(lines): break

        # Home team — appears twice
        home_team = lines[i]; i += 1
        if i < len(lines) and lines[i] == home_team:
            i += 1

        # VS separator
        if i < len(lines) and lines[i].upper().startswith("VS"):
            i += 1

        if i >= len(lines): break

        # Away team — appears twice
        away_team = lines[i]; i += 1
        if i < len(lines) and lines[i] == away_team:
            i += 1

        # GAA Logo logo line
        if i < len(lines) and "gaa logo" in lines[i].lower():
            i += 1

        if i >= len(lines): break

        # Venue — appears twice
        venue = lines[i]; i += 1
        if i < len(lines) and lines[i] == venue:
            i += 1

        # Optional referee — present only when the next line is NOT a date line
        referee = ""
        if i < len(lines) and not is_date_line(lines[i]):
            ref_raw = lines[i]; i += 1
            referee = re.sub(r"^Ref:\s*", "", ref_raw).strip()
            # skip duplicate
            if i < len(lines) and lines[i].strip() == ref_raw.strip():
                i += 1

        match_date = parse_date(date_str)
        sport      = detect_sport(competition)

        if SPORT_FILTER == "football" and sport != "Football":
            continue
        if SPORT_FILTER == "hurling" and sport != "Hurling":
            continue

        fixtures.append({
            "Date":        date_to_str(match_date) if match_date else date_str,
            "_date":       match_date,
            "Sport":       sport,
            "Competition": competition,
            "Time":        time_str,
            "Home Team":   home_team,
            "Home Score":  "",
            "Away Team":   away_team,
            "Away Score":  "",
            "Result":      "",
            "Venue":       venue,
            "Notes":       "",
        })

    print(f"  Loaded {len(fixtures)} fixtures from Manual_Fixtures file.")
    return fixtures


# ── API fetch ──────────────────────────────────────────────────────────────────

def fetch_html(action, fdate=None, tdate=None, org="3", user_id="3"):
    params = {
        "action":        action,
        "org":           org,
        "club_id":       CLUB_ID,
        "sport":         "",
        "competition_id": "",
        "team_id":       "",
        "user_id":       user_id,
        "fdate":         (fdate or date(date.today().year, 1, 1)).strftime("%Y-%m-%d"),
        "tdate":         (tdate or date.today()).strftime("%Y-%m-%d"),
        "age":           "",
        "nosuper":       "",
        "displayResults": "",
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
        "Referer":    f"https://dublingaa.sportlomo.com/clubprofile/{CLUB_ID}",
    }
    r = requests.get(API_URL, params=params, headers=headers, timeout=30)
    r.raise_for_status()
    return r.text


# ── Parse ──────────────────────────────────────────────────────────────────────

def parse_matches(html, view="results", since_date=None):
    soup = BeautifulSoup(html, "html.parser")
    matches = []

    for thead in soup.find_all("thead", class_="divider"):
        # Date
        date_div = thead.find("div", class_="date")
        date_text = date_div.get_text(strip=True) if date_div else ""
        match_date = parse_date(date_text)

        # Skip if already saved
        if since_date and match_date and match_date <= since_date:
            continue

        # Competition
        comp_link = thead.find("a")
        competition = comp_link.get_text(strip=True) if comp_link else ""
        sport = detect_sport(competition)

        # Apply sport filter
        if SPORT_FILTER == "football" and sport != "Football":
            continue
        if SPORT_FILTER == "hurling" and sport != "Hurling":
            continue

        # Each thead is followed by one or more tbody siblings
        sibling = thead.find_next_sibling()
        while sibling and sibling.name == "tbody":
            tbody = sibling
            sibling = sibling.find_next_sibling()

            desktop = tbody.find("tr", class_="desktop")
            if not desktop:
                continue

            # Time
            time_td = desktop.find("td", class_="time")
            time_text = time_td.find("span").get_text(strip=True) if time_td and time_td.find("span") else ""

            # Home team — in align-right td
            home_td = desktop.find("td", class_="align-right")
            home_name = ""
            if home_td:
                a = home_td.find("a")
                home_name = a.get_text(strip=True) if a else ""

            # Away team — in align-left td
            away_td = desktop.find("td", class_="align-left")
            away_name = ""
            if away_td:
                a = away_td.find("a")
                away_name = a.get_text(strip=True) if a else ""

            # Scores
            score_cells = desktop.find_all("td", class_="score")
            home_score = score_cells[0].get_text(strip=True) if len(score_cells) > 0 else ""
            away_score = score_cells[1].get_text(strip=True) if len(score_cells) > 1 else ""
            # Clean up empty score cells like " - "
            home_score = home_score.strip(" -") and home_score.strip() or ""
            away_score = away_score.strip(" -") and away_score.strip() or ""
            if home_score == "-" or home_score == " - ":
                home_score = ""
            if away_score == "-" or away_score == " - ":
                away_score = ""

            # Result
            result = ""
            if view == "results" and home_score and away_score:
                mh = re.match(r"(\d+)\s*-\s*(\d+)", home_score)
                ma = re.match(r"(\d+)\s*-\s*(\d+)", away_score)
                if mh and ma:
                    h_total = int(mh.group(1)) * 3 + int(mh.group(2))
                    a_total = int(ma.group(1)) * 3 + int(ma.group(2))
                    fingallians_home = "Fingallians" in home_name
                    if fingallians_home:
                        result = "W" if h_total > a_total else ("L" if h_total < a_total else "D")
                    else:
                        result = "W" if a_total > h_total else ("L" if a_total < h_total else "D")

            # Venue and notes from footer-tr
            venue = ""
            note  = ""
            footer = tbody.find("tr", class_="footer-tr")
            if footer:
                spans = footer.find_all("span")
                texts = [s.get_text(strip=True) for s in spans if s.get_text(strip=True)]
                for t in texts:
                    if "conceded" in t.lower():
                        note = t
                    elif not venue and len(t) > 2:
                        venue = t

            display_date = date_to_str(match_date) if match_date else date_text

            matches.append({
                "Date":       display_date,
                "_date":      match_date,
                "Sport":      sport,
                "Competition": competition,
                "Time":       time_text,
                "Home Team":  home_name,
                "Home Score": home_score,
                "Away Team":  away_name,
                "Away Score": away_score,
                "Result":     result,
                "Venue":      venue,
                "Notes":      note,
            })

    return matches


# ── Excel helpers ──────────────────────────────────────────────────────────────

def make_styles():
    return {
        "hf":     Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "hfill":  PatternFill("solid", start_color="003366"),
        "cf":     Font(name="Arial", size=10),
        "alt":    PatternFill("solid", start_color="EEF2F7"),
        "win":    PatternFill("solid", start_color="C6EFCE"),
        "loss":   PatternFill("solid", start_color="FFC7CE"),
        "draw":   PatternFill("solid", start_color="FFEB9C"),
        "center": Alignment(horizontal="center", vertical="center"),
        "left":   Alignment(horizontal="left",   vertical="center"),
        "border": Border(bottom=Side(style="thin", color="CCCCCC")),
    }


def write_header(ws, s):
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = s["hf"], s["hfill"], s["center"], s["border"]
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 15)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def write_header_f2(ws, s):
    for ci, col in enumerate(COLUMNS_F2, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font, cell.fill, cell.alignment, cell.border = s["hf"], s["hfill"], s["center"], s["border"]
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS_F2.get(col, 15)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def append_rows_f2(ws, data, s):
    start = ws.max_row + 1
    for ri, row in enumerate(data, start):
        for ci, col in enumerate(COLUMNS_F2, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.font, cell.border = s["cf"], s["border"]
            if col in ("Time", "Ground", "Event Type"):
                cell.alignment = s["center"]
            else:
                cell.alignment = s["left"]
                if ri % 2 == 1:
                    cell.fill = s["alt"]


def append_rows(ws, data, s):
    start = ws.max_row + 1
    for ri, row in enumerate(data, start):
        rv = row.get("Result", "")
        for ci, col in enumerate(COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
            cell.font, cell.border = s["cf"], s["border"]
            if col == "Result":
                cell.alignment = s["center"]
                if rv == "W":   cell.fill = s["win"]
                elif rv == "L": cell.fill = s["loss"]
                elif rv == "D": cell.fill = s["draw"]
            elif col in ("Home Score", "Away Score", "Time"):
                cell.alignment = s["center"]
            else:
                cell.alignment = s["left"]
                if ri % 2 == 1:
                    cell.fill = s["alt"]


def create_workbook(results, fixtures):
    wb = Workbook()
    s  = make_styles()
    ws_r = wb.active
    ws_r.title = "Results"
    write_header(ws_r, s)
    append_rows(ws_r, results, s)
    ws_f = wb.create_sheet("Fixtures")
    write_header(ws_f, s)
    append_rows(ws_f, fixtures, s)
    ws_f2 = wb.create_sheet("Fixtures_2")
    write_header_f2(ws_f2, s)
    append_rows_f2(ws_f2, build_fixtures2(fixtures), s)
    wb.save(OUTPUT_FILE)


def update_workbook(new_results, new_fixtures):
    wb = load_workbook(OUTPUT_FILE)
    s  = make_styles()

    # Results: append only new rows
    if "Results" not in wb.sheetnames:
        ws_r = wb.create_sheet("Results")
        write_header(ws_r, s)
    else:
        ws_r = wb["Results"]
    append_rows(ws_r, new_results, s)

    # Fixtures: replace entirely so changes/cancellations are reflected
    if "Fixtures" in wb.sheetnames:
        del wb["Fixtures"]
    ws_f = wb.create_sheet("Fixtures")
    write_header(ws_f, s)
    append_rows(ws_f, new_fixtures, s)

    # Fixtures_2: replace entirely (same reasoning as Fixtures)
    if "Fixtures_2" in wb.sheetnames:
        del wb["Fixtures_2"]
    ws_f2 = wb.create_sheet("Fixtures_2")
    write_header_f2(ws_f2, s)
    append_rows_f2(ws_f2, build_fixtures2(new_fixtures), s)

    wb.save(OUTPUT_FILE)


def get_last_saved_date():
    if not OUTPUT_FILE.exists():
        return None
    try:
        wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
        ws = wb["Results"] if "Results" in wb.sheetnames else None
        if not ws:
            return None
        latest = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            try:
                d = datetime.strptime(str(row[0]).strip(), "%d %B %Y").date()
                if latest is None or d > latest:
                    latest = d
            except ValueError:
                pass
        return latest
    except Exception:
        return None



# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    today      = date.today()
    last_saved = get_last_saved_date()
    first_run  = not OUTPUT_FILE.exists() or last_saved is None

    if first_run:
        fdate = date(today.year, 1, 1)
        since = None
        print(f"First run — fetching all results from 1 Jan {today.year}...")
    else:
        fdate = last_saved
        since = last_saved
        print(f"Updating — checking for results since {last_saved.strftime('%d %b %Y')}...")

    results  = []
    fixtures = []
    for org, user_id in ORGS:
        try:
            results += parse_matches(
                fetch_html("club_page_results", fdate, today, org, user_id),
                "results", since,
            )
        except Exception as e:
            print(f"  Error fetching results (org={org}): {e}")

        try:
            fixtures += parse_matches(
                fetch_html("club_page_fixtures", today, date(today.year, 12, 31), org, user_id),
                "fixtures",
            )
        except Exception as e:
            print(f"  Error fetching fixtures (org={org}): {e}")

    # Merge fixtures from Manual_Fixtures file (juvenile age grades not in API)
    fixtures += parse_manual_fixtures_file()

    # Deduplicate by (date, home team, away team, competition) in case multiple
    # orgs return overlapping data
    def dedup(rows):
        seen = set()
        out  = []
        for r in rows:
            key = (r.get("_date"), r.get("Home Team"), r.get("Away Team"), r.get("Competition"))
            if key not in seen:
                seen.add(key)
                out.append(r)
        return out

    results  = dedup(results)
    fixtures = dedup(fixtures)

    results.sort( key=lambda r: r["_date"] or date.min)
    fixtures.sort(key=lambda r: r["_date"] or date.min)

    print(f"Found {len(results)} new results, {len(fixtures)} upcoming fixtures")

    if not results and not fixtures:
        print("Nothing new to add.")
        return

    if first_run:
        create_workbook(results, fixtures)
        print(f"Created: {OUTPUT_FILE}")
    else:
        update_workbook(results, fixtures)
        print(f"Updated: {OUTPUT_FILE}")



if __name__ == "__main__":
    main()
