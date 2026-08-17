"""
Fingallians GAA – Fixtures & Results Web App
Run locally:  python3 app.py
Deploy:       push to GitHub, connect to Render (see render.yaml)
"""

import io
import os
from contextlib import redirect_stdout
from datetime import datetime
from pathlib import Path

from flask import Flask, redirect, render_template_string, url_for
from openpyxl import load_workbook

from fingallians_scraper import (
    COLUMNS,
    OUTPUT_FILE,
    main as run_scraper_main,
)

app = Flask(__name__)

LOG_FILE = Path(__file__).parent / "last_run.log"

# ── Data helpers ───────────────────────────────────────────────────────────────

def read_sheet(sheet_name):
    if not OUTPUT_FILE.exists():
        return []
    try:
        wb = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append({
                col: (str(v) if v is not None else "")
                for col, v in zip(COLUMNS, row)
            })
        wb.close()
        return rows
    except Exception as e:
        return []


def get_file_mtime():
    if OUTPUT_FILE.exists():
        ts = OUTPUT_FILE.stat().st_mtime
        return datetime.fromtimestamp(ts).strftime("%-d %b %Y, %H:%M")
    return None


def get_last_log():
    if LOG_FILE.exists():
        return LOG_FILE.read_text().strip()
    return ""


def do_run():
    buf = io.StringIO()
    try:
        with redirect_stdout(buf):
            run_scraper_main()
    except Exception as e:
        buf.write(f"\nError: {e}\n")
    LOG_FILE.write_text(buf.getvalue())


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    results  = list(reversed(read_sheet("Results")[-60:]))   # newest first
    fixtures = read_sheet("Fixtures")[:40]                    # soonest first
    return render_template_string(
        HTML_TEMPLATE,
        results=results,
        fixtures=fixtures,
        last_updated=get_file_mtime(),
        last_log=get_last_log(),
        file_exists=OUTPUT_FILE.exists(),
    )


@app.route("/run", methods=["POST"])
def run():
    do_run()
    return redirect(url_for("index"))


# ── HTML template ──────────────────────────────────────────────────────────────

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Fingallians GAA – Fixtures &amp; Results</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
         background: #eef2ee; color: #1a1a1a; min-height: 100vh; }

  /* ── Header ── */
  header {
    background: #005128; color: white;
    padding: 14px 24px;
    display: flex; align-items: center; justify-content: space-between;
    box-shadow: 0 2px 6px rgba(0,0,0,.25);
  }
  header h1 { font-size: 1.25rem; font-weight: 700; letter-spacing: -.3px; }
  header .sub { font-size: .8rem; opacity: .75; margin-top: 2px; }

  /* ── Toolbar ── */
  .toolbar {
    background: white; border-bottom: 1px solid #dde5dd;
    padding: 10px 24px;
    display: flex; align-items: center; gap: 14px; flex-wrap: wrap;
  }
  .toolbar .status { font-size: .82rem; color: #555; }

  .btn {
    background: #005128; color: white; border: none;
    padding: 9px 18px; border-radius: 6px;
    font-size: .88rem; font-weight: 600; cursor: pointer;
    display: inline-flex; align-items: center; gap: 6px;
    transition: background .15s;
  }
  .btn:hover { background: #003d1e; }
  .btn:disabled { background: #999; cursor: not-allowed; }

  /* ── Log ── */
  .log-box {
    margin: 14px 24px 0;
    background: #1e1e2e; color: #cdd6f4;
    font-family: 'SF Mono', 'Fira Mono', monospace; font-size: .78rem;
    padding: 12px 16px; border-radius: 8px; white-space: pre-wrap;
    max-height: 180px; overflow-y: auto;
  }

  /* ── Main layout ── */
  main { padding: 20px 24px; max-width: 1360px; margin: 0 auto; }
  .grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
  @media (max-width: 860px) { .grid { grid-template-columns: 1fr; } }

  /* ── Cards ── */
  .card { background: white; border-radius: 10px; overflow: hidden;
          box-shadow: 0 1px 4px rgba(0,0,0,.08); }
  .card-head {
    padding: 12px 16px; font-weight: 700; font-size: .9rem;
    display: flex; align-items: center; gap: 8px; color: white;
  }
  .card-head.green { background: #005128; }
  .card-head.gold  { background: #b07900; }
  .pill {
    background: rgba(255,255,255,.22); border-radius: 20px;
    padding: 1px 9px; font-size: .75rem; font-weight: 600;
  }

  /* ── Tables ── */
  .tbl-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: .81rem; }
  th {
    padding: 7px 10px; text-align: left;
    background: #f7f8f7; color: #555;
    font-weight: 600; font-size: .72rem; text-transform: uppercase;
    letter-spacing: .4px; border-bottom: 1px solid #eee;
  }
  td { padding: 7px 10px; border-bottom: 1px solid #f2f2f2; color: #333; vertical-align: middle; }
  tr:last-child td { border-bottom: none; }
  tr:hover td { background: #fafcfa; }
  .comp { font-size: .7rem; color: #777; margin-top: 2px; }

  /* ── Result badges ── */
  .rb {
    display: inline-block; width: 24px; height: 24px;
    border-radius: 5px; text-align: center; line-height: 24px;
    font-weight: 700; font-size: .74rem;
  }
  .rb-W { background: #c6efce; color: #1e6b1e; }
  .rb-L { background: #ffc7ce; color: #9c0006; }
  .rb-D { background: #ffeb9c; color: #7d6000; }

  /* ── Sport tag ── */
  .st { font-size: .68rem; background: #e8f0e8; color: #2d5e2d;
        border-radius: 3px; padding: 1px 5px; white-space: nowrap; }

  /* ── Score ── */
  .score { white-space: nowrap; font-variant-numeric: tabular-nums; }

  /* ── Empty / welcome ── */
  .empty { text-align: center; padding: 36px 24px; color: #888; font-size: .88rem; }
  .welcome {
    background: white; border-radius: 10px; padding: 56px 40px;
    text-align: center; box-shadow: 0 1px 4px rgba(0,0,0,.08);
  }
  .welcome h2 { color: #005128; font-size: 1.4rem; margin-bottom: 10px; }
  .welcome p  { color: #666; margin-bottom: 24px; line-height: 1.6; }

  /* ── Spinner ── */
  @keyframes spin { to { transform: rotate(360deg); } }
  .spinner {
    display: inline-block; width: 13px; height: 13px;
    border: 2px solid rgba(255,255,255,.35); border-top-color: white;
    border-radius: 50%; animation: spin .7s linear infinite;
  }
</style>
</head>
<body>

<header>
  <div>
    <h1>Fingallians GAA</h1>
    <div class="sub">Fixtures &amp; Results Tracker</div>
  </div>
</header>

<div class="toolbar">
  <button class="btn" id="run-btn" onclick="runUpdate()">▶ Run Update</button>
  {% if last_updated %}
    <span class="status">Last updated: <strong>{{ last_updated }}</strong></span>
  {% elif not file_exists %}
    <span class="status">No data yet — click Run Update to fetch from Dublin GAA</span>
  {% endif %}
</div>

{% if last_log %}
<pre class="log-box">{{ last_log }}</pre>
{% endif %}

<main>
  {% if not file_exists %}
  <div class="welcome">
    <h2>Welcome</h2>
    <p>Click <strong>Run Update</strong> above to pull all Fingallians fixtures and<br>
       results from Dublin GAA for {{ now_year }}.</p>
    <button class="btn" onclick="runUpdate()" style="font-size:1rem;padding:12px 28px">
      ▶ Run First Update
    </button>
  </div>

  {% else %}
  <div class="grid">

    <!-- Upcoming Fixtures -->
    <div class="card">
      <div class="card-head green">
        Upcoming Fixtures <span class="pill">{{ fixtures|length }}</span>
      </div>
      {% if fixtures %}
      <div class="tbl-wrap">
      <table>
        <thead>
          <tr>
            <th>Date</th>
            <th>Time</th>
            <th>Match</th>
            <th>Venue</th>
          </tr>
        </thead>
        <tbody>
        {% for f in fixtures %}
        <tr>
          <td style="white-space:nowrap">{{ f.Date }}</td>
          <td style="white-space:nowrap">{{ f.Time }}</td>
          <td>
            {{ f["Home Team"] }} <strong>v</strong> {{ f["Away Team"] }}
            <div class="comp">
              <span class="st">{{ f.Sport }}</span>
              {{ f.Competition }}
            </div>
          </td>
          <td>{{ f.Venue }}</td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      </div>
      {% else %}
      <div class="empty">No upcoming fixtures found</div>
      {% endif %}
    </div>

    <!-- Recent Results -->
    <div class="card">
      <div class="card-head gold">
        Recent Results <span class="pill">{{ results|length }}</span>
      </div>
      {% if results %}
      <div class="tbl-wrap">
      <table>
        <thead>
          <tr>
            <th>Date</th>
            <th>Match</th>
            <th>Score</th>
            <th>W/L</th>
          </tr>
        </thead>
        <tbody>
        {% for r in results %}
        <tr>
          <td style="white-space:nowrap">{{ r.Date }}</td>
          <td>
            {{ r["Home Team"] }} v {{ r["Away Team"] }}
            <div class="comp">
              <span class="st">{{ r.Sport }}</span>
              {{ r.Competition }}
            </div>
          </td>
          <td class="score">{{ r["Home Score"] }} – {{ r["Away Score"] }}</td>
          <td>
            {% if r.Result %}
            <span class="rb rb-{{ r.Result }}">{{ r.Result }}</span>
            {% endif %}
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      </div>
      {% else %}
      <div class="empty">No results found</div>
      {% endif %}
    </div>

  </div>
  {% endif %}
</main>

<script>
  async function runUpdate() {
    const btn = document.getElementById('run-btn');
    btn.disabled = true;
    btn.innerHTML = '<span class="spinner"></span> Updating…';

    try {
      const res = await fetch('/run', { method: 'POST' });
      if (res.ok || res.redirected) {
        window.location.reload();
      }
    } catch (e) {
      alert('Update failed: ' + e);
      btn.disabled = false;
      btn.innerHTML = '▶ Run Update';
    }
  }
</script>

</body>
</html>
"""


# ── Entry point ────────────────────────────────────────────────────────────────

# Make current year available to the template
@app.context_processor
def inject_globals():
    return {"now_year": datetime.now().year}


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
